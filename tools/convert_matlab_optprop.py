"""Convert the lab's MATLAB optprop_* material routines into meent nk_data text tables.

The MATLAB library under 'Material data' exposes one optprop_<material>_<source>.m per dataset.
Three shapes appear, and this script covers all of them:

  * thin wrappers that load 'wavnk_<name>.mat', whose 'data' variable is an (N, 3) array of
    [wavelength(um), n, k];
  * inline oscillator models (optprop_Au, optprop_Ti, optprop_CaF2);
  * a parametric Kubo model (optprop_Gr_Falkovsky), which is a model family rather than one
    material -- graphene's response depends on how it is gated, so a table has to fix the Fermi
    level, mobility and temperature.

Sources are located by filename anywhere under the directories given, so it does not matter
whether the library is laid out in subfolders or flat; anything missing is reported and skipped.

Measurements are written out as data rows. The inline models are written out as coefficients
instead, under a 'type: model <name>' header that meent/dispersion.py evaluates on lookup -- no
resampling, and the parameters stay visible and editable. Graphene is the exception: its Kubo
integral needs quadrature per point, so it stays sampled.

Output goes to nk_data/jLab/. Tables whose upstream is the refractiveindex.info database
belong in nk_data/refractiveindex_info/ via tools/convert_refractiveindex_info.py instead --
Ti_BrendelBormann is one such case, being byte-identical to that database's Ti/Rakic-BB.

Usage:
    python tools/convert_matlab_optprop.py "<path to the material data directory>"
"""

import argparse
from datetime import date
from pathlib import Path

import numpy as np
from scipy.integrate import quad
from scipy.io import loadmat
from scipy.special import wofz

# Physical constants, matching the values hard-coded in the MATLAB sources so the ported models
# reproduce them exactly.
ELEMENTARY_CHARGE = 1.6021773349e-19
BOLTZMANN = 1.38064852e-23
HBAR = 1.0545718e-34
VACUUM_PERMITTIVITY = 8.85418781762039e-12
FERMI_VELOCITY = 1e6
LIGHT_SPEED = 299792458

# (output name, wavnk_*.mat stem) -- wrappers that just load a [wavelength(um), n, k] table.
# Output files are named after their source: same stem, .txt extension. That is also the lookup
# key, since read_material_table keys the table on the file name.
#
# (material, source .mat) -- wrappers that just load a [wavelength(um), n, k] table.
MAT_TABLES = [
    ('MgO', 'wavnk_MgO_Palik.mat'),
    ('Si', 'wavnk_Si_JwKang_260409.mat'),
    ('SiO2', 'wavnk_SiO2_JwKang_260409.mat'),
]

# (material, source .xlsx) -- [wavelength(um), n_o, k_o, n_e, k_e] under one header row, for
# uniaxial crystals. One file holds one index, so these are written out as separate -o and -e
# files: a single file would lose the extraordinary ray silently, since find_nk_index reads
# columns 1 and 2.
XLSX_BIREFRINGENT = [
    ('BaTiO3', '260223_BTO_Vis_MIR_intrinsic.xlsx'),
]

# Fermi levels tabulated for the Falkovsky model, in eV. Graphene is transparent below the
# interband onset at hw = 2 Ef, so each level gives a materially different curve.
GRAPHENE_SOURCE = 'optprop_Gr_Falkovsky.m'
GRAPHENE_FERMI_LEVELS = [0.2, 0.4, 0.6]
GRAPHENE_MOBILITY = 10000.0  # cm^2/Vs
GRAPHENE_TEMPERATURE = 300.0  # K
GRAPHENE_THICKNESS = 0.34e-9  # m, used to turn a sheet conductivity into a bulk permittivity

# Wavelength grid for the sampled model, log spaced because it spans two decades.
MODEL_WAVELENGTH_RANGE = (0.5e-6, 50e-6)
MODEL_SAMPLES = 1000

# The optprop_* routines below compute a Brendel-Bormann oscillator model inline instead of
# loading a table: a Drude term for the conduction electrons plus Gaussian-broadened Lorentz
# oscillators for the interband transitions, which is why it needs the Faddeeva function.
#
# These are kept as coefficients and evaluated on lookup, like the Sellmeier fits under
# refractiveindex_info/. Those evaluate with plain arithmetic in whichever array library a backend
# uses; this one needs wofz, so it goes through numpy -- which is what every backend but jax
# hands to dispersion.evaluate anyway.
#
# The coefficients are the L vectors the .m files hard-code; the sources are read only to confirm
# the dataset is the one being converted.
# (material, source .m, parameter convention, coefficients)
#   'ev' -- optprop_Au.m / optprop_Ti.m, which convert cm-1 to eV first
#   'cm' -- optprop_CaF2.m, which works in cm-1 throughout
BRENDEL_BORMANN = [
    ('Au', 'optprop_Au.m', 'ev',
     [9.20007, 0.85195, 0.09509, 0.03361, 0.08138, 0.24435, 0.66983, 0.04378,
      0.02654, 3.57689, 0.30055, 0.39482, 0.09543, 3.59156, 0.93598, 0.83641,
      7.69005, 3.59644, 0.88825, 1.28384, 0.23893, 34.87059, 1.87353]),
    ('Ti', 'optprop_Ti.m', 'ev',
     [7.29, 0.126, 0.067, 0.427, 1.877, 1.459, 0.463, 0.218, 0.1, 2.661, 0.506,
      0.513, 0.615, 0.805, 0.799, 0.0002, 4.109, 19.86, 2.854]),
    ('CaF2', 'optprop_CaF2.m', 'cm',
     [1.84462, 14357190.64106, 2.59652, 626.33865, 11858.60141, 269739.04488,
      0.14187, 223.90971, 53.62512, 786.56174, 33.25617, 242.183, 504.21943]),
]
BRENDEL_BORMANN_RANGE = (0.15e-6, 100e-6)
BRENDEL_BORMANN_SAMPLES = 4000

# (output name, .mat file, variable) -- tables in the newer set, stored as [wavenumber(cm-1), n, k]
# rather than the [wavelength(um), n, k] the older library uses.
WAVENUMBER_TABLES = [
    ('Iongel', '[Optprop][Iongel][JW].mat', 'Iongel_JW'),
]


def _cosh_ratio(a, b):
    """cosh(a) / cosh(b) for positive arguments, without overflowing at large b."""
    return np.exp(a - b) * (1 + np.exp(-2 * a)) / (1 + np.exp(-2 * b))


def graphene_falkovsky(wavelength, fermi_level, mobility, temperature):
    """Sheet conductivity of graphene from the Falkovsky/Kubo model, ported from MATLAB.

    Returns the complex refractive index obtained by spreading the sheet conductivity over
    GRAPHENE_THICKNESS, which is the convention the MATLAB routine uses.
    """
    temperature_ev = temperature * BOLTZMANN / ELEMENTARY_CHARGE
    mobility_si = mobility * 1e-4
    photon_ev = HBAR * (2 * np.pi * LIGHT_SPEED / wavelength) / ELEMENTARY_CHARGE

    # Impurity-limited scattering rate, in eV.
    scattering = abs(HBAR * FERMI_VELOCITY ** 2
                     / (mobility_si * ELEMENTARY_CHARGE * fermi_level))

    def occupation(energy):
        return (np.tanh(energy / temperature_ev)
                / (_cosh_ratio(fermi_level / temperature_ev, energy / temperature_ev) + 1.0))

    conductivity = np.empty(photon_ev.shape, dtype=complex)
    for index, energy in enumerate(photon_ev):
        half = occupation(energy / 2)

        # The pole at z = energy/2 is removable because the numerator vanishes there; quad still
        # needs to be told about it to place its panels sensibly.
        integrand = lambda z: (occupation(z) - half) / (energy ** 2 - 4 * z * z)
        interband, _ = quad(integrand, 0, 1000, points=[energy / 2], limit=400)

        intraband = ((2j * ELEMENTARY_CHARGE ** 2 * temperature_ev)
                     / (np.pi * HBAR * (energy + 1j * scattering))
                     * np.log(2 * np.cosh(fermi_level / (2 * temperature_ev))))
        conductivity[index] = intraband + (ELEMENTARY_CHARGE ** 2 / (4 * HBAR)) * (
            half + (4j * energy / np.pi) * interband)

    angular_frequency = ELEMENTARY_CHARGE * photon_ev / HBAR
    permittivity = 1 + 1j * conductivity / (
        angular_frequency * GRAPHENE_THICKNESS * VACUUM_PERMITTIVITY)
    return np.sqrt(permittivity)


def brendel_bormann(coefficients, wavenumber, convention):
    """Brendel-Bormann permittivity, ported from optprop_Au.m / optprop_Ti.m / optprop_CaF2.m.

    A Lorentz oscillator broadened by a Gaussian spread in its centre frequency integrates to a
    Voigt profile, which is where the Faddeeva function comes from. The two conventions differ
    only in whether the parameters are quoted in eV or in wavenumbers.
    """
    if convention == 'ev':
        omega = np.asarray(wavenumber, dtype=float) / 8065.54429  # cm-1 per eV
        plasma, strength, damping = coefficients[:3]
        permittivity = 1 - strength * plasma ** 2 / (omega * (omega + 1j * damping))
        for i in range(3, len(coefficients), 4):
            f, gamma, centre, sigma = coefficients[i:i + 4]
            a = np.sqrt(omega ** 2 + 1j * omega * gamma)
            permittivity = permittivity + 1j * np.sqrt(np.pi) * f * plasma ** 2 / (
                2 ** 1.5 * a * sigma) * (wofz((a - centre) / (np.sqrt(2) * sigma))
                                         + wofz((a + centre) / (np.sqrt(2) * sigma)))
        return np.sqrt(permittivity)

    omega = np.asarray(wavenumber, dtype=float)
    permittivity = coefficients[0]
    for i in range(1, len(coefficients), 4):
        strength, gamma, sigma, centre = coefficients[i:i + 4]
        a = np.sqrt(omega ** 2 + 1j * gamma * omega)
        permittivity = permittivity + 1j * np.sqrt(np.pi) * strength / (
            2 * np.sqrt(2) * sigma * a) * (wofz((a - centre) / (np.sqrt(2) * sigma))
                                           + wofz((a + centre) / (np.sqrt(2) * sigma)))
    return np.sqrt(permittivity)


def write_table(out_dir, name, wavelength, n, k, header):
    lines = ['Wavelength(m)\tn\tk']
    lines += [f'# {line}' for line in header]
    lines.append(f'# range: {wavelength[0]:.4e} - {wavelength[-1]:.4e} m')
    lines.append(f'# converted: {date.today().isoformat()} by tools/convert_matlab_optprop.py')
    for wavelength_value, n_value, k_value in zip(wavelength, n, k):
        lines.append(f'{wavelength_value:.6e}\t{n_value:.6g}\t{k_value:.6g}')

    path = Path(out_dir) / f'{name}.txt'
    path.write_text('\n'.join(lines) + '\n', encoding='ascii')
    return path, len(wavelength)


def read_spreadsheet(path):
    """[wavelength(um), n_o, k_o, n_e, k_e] from a one-sheet .xlsx with a single header row."""
    from openpyxl import load_workbook

    sheet = load_workbook(str(path), data_only=True, read_only=True).worksheets[0]
    rows = sheet.iter_rows(min_row=2, values_only=True)
    return np.array([row[:5] for row in rows if row[0] is not None], dtype=float)


def write_birefringent(out_dir, name, material, table, source):
    """Write [wavelength(um), n_o, k_o, n_e, k_e] out as separate -o and -e tables."""
    wavelength = table[:, 0] * 1e-6
    for ray, label, n_col, k_col in [('o', 'ordinary', 1, 2), ('e', 'extraordinary', 3, 4)]:
        path, count = write_table(
            out_dir, f'{name}-{ray}', wavelength, table[:, n_col], table[:, k_col],
            [f'material: {material}',
             f'source: {source}',
             f'ray: {label}',
             'note: uniaxial crystal; pair with the matching -o / -e table to build a '
             '(nx, ny, nz) ucell',
             'type: tabulated nk'])
        print(f'{path.name:38} {count:5} pts')


def write_model(out_dir, name, model, coefficients, wavelength_range, header):
    """Write a model as its coefficients, with no data rows.

    meent/dispersion.py reads the 'type: model' and 'coefficients' lines back and evaluates the
    curve at lookup time, so nothing here is resampled and the parameters stay editable.
    """
    lines = ['Wavelength(m)\tn\tk']
    lines += [f'# {line}' for line in header]
    lines.append(f'# type: model {model}')
    lines.append('# coefficients: ' + ' '.join(repr(float(value)) for value in coefficients))
    lines.append(f'# range: {wavelength_range[0]:.4e} - {wavelength_range[1]:.4e} m')
    lines.append(f'# converted: {date.today().isoformat()} by tools/convert_matlab_optprop.py')

    path = Path(out_dir) / f'{name}.txt'
    path.write_text('\n'.join(lines) + '\n', encoding='ascii')
    return path, len(coefficients)


def find_source(roots, filename):
    """First file called `filename` anywhere under `roots`, or None.

    Searching by name rather than by path keeps this working whether the library is laid out in
    subfolders or flat, which the two sets it has to read are not consistent about.
    """
    for root in roots:
        if root is None:
            continue
        candidate = Path(root) / filename
        if candidate.is_file():
            return candidate
        for match in Path(root).rglob(filename):
            return match
    return None


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('material_data', help='path to the material data directory')
    parser.add_argument('--jw-dir', help='second directory to search, if the sets are split')
    parser.add_argument('-o', '--out-dir',
                        default=str(Path(__file__).resolve().parent.parent
                                    / 'meent' / 'nk_data' / 'jLab'))
    args = parser.parse_args()

    roots = [args.material_data, args.jw_dir]
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    missing = []

    for material, filename in MAT_TABLES:
        source = find_source(roots, filename)
        if source is None:
            missing.append(filename)
            continue
        table = np.asarray(loadmat(str(source))['data'], dtype=float)
        wavelength = table[:, 0] * 1e-6  # the MATLAB library tabulates micrometres
        path, count = write_table(
            out_dir, Path(filename).stem, wavelength, table[:, 1], table[:, 2],
            [f'material: {material}',
             f'source: {filename}',
             'type: tabulated nk'])
        print(f'{path.name:38} {count:5} pts')

    # A uniaxial crystal needs both rays, and find_nk_index reads one (n, k) pair per file, so
    # each ray is written out separately. They are looked up as two independent materials and
    # combined by the caller -- meent has no notion of an optic axis.
    for material, filename in XLSX_BIREFRINGENT:
        source = find_source(roots, filename)
        if source is None:
            missing.append(filename)
            continue
        write_birefringent(out_dir, Path(filename).stem, material,
                           read_spreadsheet(source), filename)

    # Inline models: coefficients only, evaluated by meent/dispersion.py on lookup.
    for material, filename, convention, coefficients in BRENDEL_BORMANN:
        if find_source(roots, filename) is None:
            missing.append(filename)
            continue
        unit = 'eV' if convention == 'ev' else 'cm-1'
        layout = ('[wp, f0, g0, (f, gamma, centre, sigma) * n]' if convention == 'ev'
                  else '[eps_inf, (strength, gamma, sigma, centre) * n]')
        path, count = write_model(
            out_dir, Path(filename).stem, f'brendel-bormann-{convention}', coefficients,
            BRENDEL_BORMANN_RANGE,
            [f'material: {material}',
             f'source: {filename}',
             'model: Brendel-Bormann oscillators (Drude term plus Gaussian-broadened Lorentz '
             'oscillators, hence the Faddeeva function)',
             f'coefficient layout: {layout}, in {unit}'])
        print(f'{path.name:38} {count:5} coefficients')

    for material, filename, variable in WAVENUMBER_TABLES:
        source = find_source(roots, filename)
        if source is None:
            missing.append(filename)
            continue
        table = np.asarray(loadmat(str(source))[variable], dtype=float)
        # Stored against wavenumber, so wavelength runs backwards; interp needs it ascending.
        table = table[np.argsort(1e-2 / table[:, 0])]
        path, count = write_table(
            out_dir, Path(filename).stem, 1e-2 / table[:, 0], table[:, 1], table[:, 2],
            [f'material: {material}',
             f'source: {filename}',
             'type: tabulated nk'])
        print(f'{path.name:38} {count:5} pts')

    # Graphene stays sampled: the Kubo interband term is an integral evaluated by quadrature per
    # point, which is not something a coefficient line can carry. One source, one file per Fermi
    # level, so the level is the only thing appended to the name.
    if find_source(roots, GRAPHENE_SOURCE) is None:
        missing.append(GRAPHENE_SOURCE)
    else:
        wavelength = np.geomspace(*MODEL_WAVELENGTH_RANGE, MODEL_SAMPLES)
        for fermi_level in GRAPHENE_FERMI_LEVELS:
            index = graphene_falkovsky(wavelength, fermi_level,
                                       GRAPHENE_MOBILITY, GRAPHENE_TEMPERATURE)
            name = f'{Path(GRAPHENE_SOURCE).stem}-ef{round(fermi_level * 1000)}meV'
            path, count = write_table(
                out_dir, name, wavelength, index.real, index.imag,
                ['material: Graphene',
                 f'source: {GRAPHENE_SOURCE} (Falkovsky/Kubo model)',
                 f'conditions: Ef = {fermi_level} eV, mobility = {GRAPHENE_MOBILITY:g} cm^2/Vs, '
                 f'T = {GRAPHENE_TEMPERATURE:g} K',
                 f'note: sheet conductivity spread over {GRAPHENE_THICKNESS:g} m to give a bulk '
                 'index; valid only at this gating',
                 'type: model'])
            print(f'{path.name:38} {count:5} pts')

    if missing:
        print('\nnot found, skipped:')
        for filename in missing:
            print(f'  {filename}')


if __name__ == '__main__':
    main()
